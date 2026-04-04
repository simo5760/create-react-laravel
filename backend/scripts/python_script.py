import os
import shutil
import re
import logging
from datetime import datetime
from PyPDF2 import PdfReader
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

import pytesseract
from pdf2image import convert_from_path


# =========================
# CONFIGURATION
# =========================

INPUT_FOLDER = "backend/storage/app/uploads"
OUTPUT_FOLDER = "backend/storage/app/output"
BACKUP_FOLDER = "backend/storage/app/backup"
TEMP_FOLDER = "backend/temp_images"
EXCEL_FILE = "backend/storage/app/Excel/documents_classes.xlsx"
LOG_FILE = "backend/storage/app/log.txt"

POPPLER_PATH = r"C:\Users\moham\Downloads\Release-25.12.0-0\poppler-25.12.0\Library\bin"
pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

CATEGORIES = {
     "A_003": [
"AUTORISATION DU PROPRIETAIRE",
"AUTORISATION DU PROPRIÉTAIRE",
"AUTORISATION PROPRIETAIRE",
"AUTORISATION DE PROPRIETAIRE",
"AUTORISATION D'ABONNEMENT",
"PROPRIETAIRE DE L'IMMEUBLE",
"AU NOM ET POUR LE COMPTE DE",
"GERANT AGISSANT AU NOM",
"CONTRAT D'ABONNEMENT",
"FOURNITURE D'EAU",
"FOURNITURE D'ELECTRICITE",
"IMMEUBLE SIS A",
"LOCATAIRE",
"ABONNEMENT",
"COMPTEUR EAU",
"COMPTEUR ELECTRICITE",

"رخصة الملاك",
"رخصة ملاك",
"إذن الملاك",
"تصريح الملاك",
"مالك العقار",
"مالك العمارة",
"عقد الاشتراك",
"الاشتراك",
"التزويد بالماء",
"التزويد بالكهرباء",
"المكتري",
"الساكن",
"الشقة",
"باسم ولحساب",

":السيد (ة)",
":رقم الجولة",
"radeet",
"beni-mellal"
],
     "A_006": ["n° de police", "رقم الوثيقة"],
     "A_007": ["contrat de bail", "عقد كراء"],
     "A_013": ["facture", "فاتورة"],
     "A_016": ["procuration", "وكالة", "موافقة"],
     "A_017": ["reçu", "n° recu"],
     "I_004": ["carte nationale d'identite", "royaume du maroc","البطاقة الوطنية للتعريف "]

}

DEFAULT_CATEGORY = "A_004"


# =========================
# LOGGING
# =========================

logging.basicConfig(
    filename=LOG_FILE,
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s"
)


# =========================
# CREATION DOSSIERS
# =========================

def create_folders():
    os.makedirs(INPUT_FOLDER, exist_ok=True)
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)
    os.makedirs(BACKUP_FOLDER, exist_ok=True)
    os.makedirs(TEMP_FOLDER, exist_ok=True)
    os.makedirs("excel", exist_ok=True)

    for category in CATEGORIES.keys():
        os.makedirs(os.path.join(OUTPUT_FOLDER, category), exist_ok=True)

    os.makedirs(os.path.join(OUTPUT_FOLDER, DEFAULT_CATEGORY), exist_ok=True)


# =========================
# OCR PDF
# =========================

def ocr_pdf(pdf_path):

    text = ""

    try:
        images = convert_from_path(pdf_path, poppler_path=POPPLER_PATH)

        for i, image in enumerate(images):
            temp_image_path = os.path.join(TEMP_FOLDER, f"page_{i}.png")

            image.save(temp_image_path, "PNG")

            text += pytesseract.image_to_string(
                temp_image_path,
                lang="fra+eng+ara",
                config="--oem 3 --psm 6"
            )

            os.remove(temp_image_path)

    except Exception as e:
        logging.error(f"OCR error: {pdf_path} - {e}")
        print(f"❌ Erreur OCR avec {pdf_path}")

    return text.lower()


# =========================
# EXTRACTION TEXTE
# =========================

def extract_text_from_pdf(pdf_path):

    text = ""

    try:
        reader = PdfReader(pdf_path)

        for page in reader.pages:
            page_text = page.extract_text()
            if page_text:
                text += page_text

    except Exception as e:
        logging.warning(f"Lecture PDF échouée: {pdf_path} - {e}")

    if not text.strip():
        text = ocr_pdf(pdf_path)

    return text.lower()


# =========================
# CORRECTION OCR
# =========================

def clean_text(text):

    text = text.lower()

    corrections = {
       "prorietaire": "proprietaire",
        "n?": "n°",
        "pol1ce": "police",
        "pollce": "police",
       
        "recü": "reçu",
        "recu": "reçu",
        "cin ": "carte nationale d'identite",
        "رقم الوثيقه": "رقم الوثيقة",
        "عقد كرا": "عقد كراء"
    }

    for wrong, correct in corrections.items():
        text = text.replace(wrong, correct)

    return text


# # =========================
# # EXTRAIRE NUMERO
# # =========================
#
def extract_number(text):

    patterns = [
        r"n°\s*de\s*police\s*[:\-]?\s*(\d+)",
        r"رقم الوثيقة\s*[:\-]?\s*(\d+)"
    ]

    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            return match.group(1)

    return ""


# =========================
# CLASSIFICATION INTELLIGENTE
# =========================

def classify_document(text):

    text = text.lower()

    # PRIORITÉ

    if "رقم الوثيقة" in text or "n° de police" in text:
        return "A_006"

    if "contrat de bail" in text or "عقد كراء" in text:
        return "A_007"

    if "facture" in text or "فاتورة" in text:
        return "A_013"

    if "procuration" in text or "وكالة" in text or  "موافقة"in text:
        return "A_016"

    if "reçu" in text or "n° recu" in text :
        return "A_017"

    if "carte nationale d'identite" in text or "royaume du maroc" in text or "البطاقة الوطنية للتعريف " in text :
        return "I_004"









    return DEFAULT_CATEGORY


# =========================
# ANALYSE DOCUMENTS
# =========================

def analyze_documents():

    results = []

    for filename in os.listdir(INPUT_FOLDER):

        if filename.lower().endswith(".pdf"):

            pdf_path = os.path.join(INPUT_FOLDER, filename)

            try:
                text = clean_text(extract_text_from_pdf(pdf_path))

                category = classify_document(text)

                # Numero seulement pour A_006
                number = extract_number(text) if category == "A_006" else ""

                destination_folder = os.path.join(OUTPUT_FOLDER, category)
                new_path = os.path.join(destination_folder, filename)
                backup_path = os.path.join(BACKUP_FOLDER, filename)

                shutil.copy(pdf_path, backup_path)
                shutil.move(pdf_path, new_path)

                date_now = datetime.now().strftime("%Y-%m-%d %H:%M")

                results.append(
                    (filename, os.path.abspath(new_path), date_now, category, number)
                )

                logging.info(f"Succès: {filename}")

            except Exception as e:

                logging.error(f"Erreur avec {filename}: {e}")

                choice = input(f"Erreur avec {filename}. Continuer ? (y/n): ")

                if choice.lower() != "y":
                    break

    return results


# =========================
# EXPORT EXCEL
# =========================

def export_to_excel(data):

    wb = Workbook()
    ws = wb.active
    ws.title = "Documents"

    headers = [
        "Nom Document",
        "Chemin",
        "Date",
        "Type",
        "Numero"
    ]

    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    colors = {
        "A_013": "FFF2CC",
        "A_007": "D9EAD3",
        "I_004": "CFE2F3",
        "A_006": "F4CCCC"
    }

    for row_idx, (name, path, date, doc_type, number) in enumerate(data, start=2):

        ws.cell(row=row_idx, column=1).value = name

        link_cell = ws.cell(row=row_idx, column=2)
        link_cell.value = path
        link_cell.hyperlink = f"file:///{path.replace(os.sep, '/')}"
        link_cell.style = "Hyperlink"

        ws.cell(row=row_idx, column=3).value = date

        type_cell = ws.cell(row=row_idx, column=4)
        type_cell.value = doc_type

        if doc_type in colors:
            type_cell.fill = PatternFill(
                start_color=colors[doc_type],
                end_color=colors[doc_type],
                fill_type="solid"
            )

        ws.cell(row=row_idx, column=5).value = number

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = max_length + 3

    ws.auto_filter.ref = ws.dimensions

    wb.save(EXCEL_FILE)

    print("Excel généré :", EXCEL_FILE)


# =========================
# MAIN
# =========================

def main():

    print("Création des dossiers...")
    create_folders()

    print("Analyse des documents...")
    results = analyze_documents()

    print("Création Excel...")
    export_to_excel(results)

    print("Terminé ✔")


if __name__ == "__main__":
    main()